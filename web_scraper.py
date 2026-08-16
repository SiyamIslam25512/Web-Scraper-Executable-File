import time
from datetime import timedelta, date
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, urlparse, urldefrag
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os

def main():
    url = input("Website URL: ").strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    while True:
        start_date = input("Enter Start Date (YYYY-MM-DD): ").strip()
        end_date = input("Enter End Date (YYYY-MM-DD): ").strip()
        try:
            start_date = date.fromisoformat(start_date)
            end_date = date.fromisoformat(end_date)
            if start_date > end_date:
                print("Invalid Range: Start date cannot be after End date. Please try again")
                continue
            print(f"Fetching data from {start_date} to {end_date}")
            break
        except ValueError:
            print("Incorrect format! Please use YYYY-MM-DD format (e.g., 2026-08-01).")
            continue

    date_list = generate_date_list(start_date, end_date)

    # ১) পুরো সাইট ক্রল
    discovered_pages = crawl_website(url)
    print(f"Found {len(discovered_pages)} HTML pages.")

    # মূল URL না থাকলে যোগ করা
    if url.rstrip("/") not in [p.rstrip("/") for p in discovered_pages]:
        discovered_pages.insert(0, url)

    # ২) প্রতিটা পেজ × প্রতিটা তারিখ
    all_data, scraped_urls = scrape_pages_with_dates(discovered_pages, date_list)

    all_data = remove_duplicate_datasets(all_data)
    all_data = merge_datasets_by_title(all_data)

    fill_date = start_date.strftime("%Y-%m-%d")
    for dataset in all_data:
        for row in dataset.get("rows", []):
            if not row.get("Date"):
                row["Date"] = fill_date

    # ===== Excel file name =====
    default_name = f"scraped_data_{date.today().strftime('%Y%m%d')}_{time.strftime('%H%M%S')}"

    while True:
        user_name = input(
            f"\nEnter a name for the Excel file (press Enter to use the default name '{default_name}')\n"
            f"Please do not type .xlsx — it will be added automatically: "
        ).strip()

        if not user_name:
            filename = default_name + ".xlsx"
        else:
            if user_name.lower().endswith(".xlsx"):
                user_name = user_name[:-5].strip()

            if not user_name:
                print("Invalid name. Please type a proper file name without .xlsx")
                continue

            filename = user_name + ".xlsx"

        if os.path.exists(filename):
            print(f"\nA file named '{filename}' already exists.")
            choice = input(
                "Do you want to replace the old file with the new data?\n"
                "Type 'yes' or 'y' to replace, and 'no' or 'n' to choose a different name: "
            ).strip().lower()

            if choice in ("yes", "y", "Yes", "YES", "YEs", "YeS", "yES"):
                break
            elif choice in ("no", "n", "No", "NO", "nO"):
                print("Okay, please enter a different file name.\n")
                continue
            else:
                print("Please type only 'yes' or 'y' to replace, or only 'no' or 'n' to choose a different name.\n")
                continue
        else:
            break

    save_to_excel(all_data, filename)

def build_date_urls(base_url, formatted_date):

    possible_urls = []

    # 1. Placeholder style
    if "{date}" in base_url:
        possible_urls.append(base_url.replace("{date}", formatted_date))

    # 2. Common query parameter styles
    if "?" in base_url:
        possible_urls.append(f"{base_url}&date={formatted_date}")
        possible_urls.append(f"{base_url}&d={formatted_date}")
        possible_urls.append(f"{base_url}&start_date={formatted_date}")
        possible_urls.append(f"{base_url}&day={formatted_date}")
    else:
        possible_urls.append(f"{base_url}?date={formatted_date}")
        possible_urls.append(f"{base_url}?d={formatted_date}")
        possible_urls.append(f"{base_url}?start_date={formatted_date}")
        possible_urls.append(f"{base_url}?day={formatted_date}")

    # 3. Path style (example.com/2026-08-01/)
    if base_url.endswith("/"):
        possible_urls.append(f"{base_url}{formatted_date}/")
    else:
        possible_urls.append(f"{base_url}/{formatted_date}/")
        possible_urls.append(f"{base_url}/{formatted_date}")

    # Remove duplicates while keeping order
    unique_urls = list(dict.fromkeys(possible_urls))
    return unique_urls

def generate_date_list(start_date, end_date):
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    return date_list

def sending_request_to_web(date_list, url):
    all_data = []
    scraped_urls = set()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 11.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"
    }

    # Selenium driver একবারই খোলা
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--log-level=3")

    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        for target_date in date_list:
            formatted_date = target_date.strftime("%Y-%m-%d")
            print(f"\nScraping data for: {formatted_date}...")

            possible_urls = build_date_urls(url, formatted_date)
            found_any = False

            for target_url in possible_urls:
                print(f"  Trying: {target_url}")
                page_results = []

                # ১) requests দিয়ে স্ট্যাটিক ডেটা
                try:
                    response = requests.get(target_url, headers=headers, timeout=10)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, "html.parser")
                        page_data = extract_data(soup)
                        if page_data:
                            page_results.extend(page_data)
                            print("    → Static data found with requests")
                except requests.exceptions.RequestException:
                    pass

                # ২) Selenium দিয়ে JavaScript ডেটা
                try:
                    driver.get(target_url)
                    time.sleep(3)
                    soup = BeautifulSoup(driver.page_source, "html.parser")
                    page_data = extract_data(soup)
                    if page_data:
                        page_results.extend(page_data)
                        print("    → JavaScript data found with Selenium")
                except Exception as e:
                    print(f"    → Selenium error: {e}")

                # ডেটা পেলে Date বসিয়ে সেভ
                if page_results:
                    for dataset in page_results:
                        for row in dataset.get("rows", []):
                            row["Date"] = formatted_date
                    all_data.extend(page_results)
                    scraped_urls.add(target_url.rstrip("/"))
                    found_any = True
                    print(f"  → Success with this URL")
                    break  # এই তারিখের জন্য একটা সফল URL পেলেই যথেষ্ট

            if not found_any:
                print(f"  → No data found for {formatted_date}")

            time.sleep(1)

    finally:
        if driver:
            driver.quit()

    return all_data, scraped_urls

def extract_data(soup):
    data = []
    tables = soup.find_all("table")

    for table in tables:
        rows = table.find_all("tr")
        if not rows:
            continue

        raw_headers = [cell.get_text(" ", strip=True) for cell in rows[0].find_all(["th", "td"])]
        if not raw_headers:
            continue

        # ডুপ্লিকেট হেডার ইউনিক করা
        headers = []
        seen = {}
        for h in raw_headers:
            if h == "":
                h = "Column"
            if h in seen:
                seen[h] += 1
                headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 1
                headers.append(h)

        data_rows = []
        for row in rows[1:]:
            cells = row.find_all(["th", "td"])
            values = [cell.get_text(" ", strip=True) for cell in cells]

            if values:
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(values):
                        row_dict[header] = values[i]
                    else:
                        row_dict[header] = ""
                data_rows.append(row_dict)

        if data_rows:
            # টেবিলের আগের heading খোঁজা
            title = "Data Table"
            prev = table.find_previous(["h1", "h2", "h3", "h4"])
            if prev:
                title = prev.get_text(" ", strip=True)

            data.append({
                "title": title,
                "headers": headers,
                "rows": data_rows
            })

    return data

def crawl_website(start_url, headers=None):
    if headers is None:
        headers = {"User-Agent": ("Mozilla/5.0 (Windows NT 11.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36")}

    start_url = urldefrag(start_url)[0]

    parsed_start_url = urlparse(start_url)

    if parsed_start_url.scheme not in ("http", "https"):
        print("Invalid URL.")
        return []

    base_domain = parsed_start_url.netloc.lower()

    urls_to_visit = [start_url]

    visited_urls = set()

    discovered_pages = []

    while urls_to_visit:

        current_url = urls_to_visit.pop(0)

        current_url = urldefrag(current_url)[0]

        if current_url in visited_urls:
            continue

        parsed_url = urlparse(current_url)

        if parsed_url.scheme not in ("http", "https"):
            continue

        if parsed_url.netloc.lower() != base_domain:
            continue

        visited_urls.add(current_url)

        print(f"Crawling: {current_url}")

        try:
            response = requests.get(current_url,headers=headers,timeout=10)

            if response.status_code != 200:
                print(f"Could not access {current_url}: HTTP {response.status_code}")
                continue

            content_type = response.headers.get("Content-Type","").lower()

            if "text/html" not in content_type:
                continue

            discovered_pages.append(current_url)

            soup = BeautifulSoup(response.text,"html.parser")

            for link in soup.find_all("a", href=True):

                href = link["href"].strip()

                if not href:
                    continue

                if href.startswith((
                    "#",
                    "mailto:",
                    "tel:",
                    "javascript:",
                    "data:")):
                    continue

                absolute_url = urljoin(current_url,href)

                absolute_url = urldefrag(absolute_url)[0]

                parsed_link = urlparse(absolute_url)

                if parsed_link.scheme not in ("http", "https"):
                    continue

                if parsed_link.netloc.lower() != base_domain:
                    continue

                if absolute_url not in visited_urls:
                    urls_to_visit.append(absolute_url)

            time.sleep(0.5)

        except requests.exceptions.RequestException as e:
            print(f"Network error while crawling {current_url}: {e}")

    return discovered_pages

def extract_javascript_data(discovered_pages):
    javascript_data = []

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--log-level=3")

    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        for page_url in discovered_pages:
            print(f"Checking page: {page_url}")
            page_results = []

            # ১) আগে requests দিয়ে স্ট্যাটিক ডেটা নাও
            try:
                headers = {
                    "User-Agent": "Mozilla/5.0 (Windows NT 11.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"
                }
                response = requests.get(page_url, headers=headers, timeout=10)

                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, "html.parser")
                    page_data = extract_data(soup)

                    if page_data:
                        for dataset in page_data:
                            for row in dataset.get("rows", []):
                                if "Date" not in row:
                                    row["Date"] = ""
                        page_results.extend(page_data)
                        print("  → Static data found with requests")
            except Exception as e:
                print(f"  → requests error: {e}")

            # ২) তারপর Selenium দিয়ে JavaScript ডেটাও নাও
            try:
                print(f"  → Rendering JavaScript: {page_url}")
                driver.get(page_url)
                time.sleep(3)  # JS লোড হওয়ার সময়

                rendered_html = driver.page_source
                soup = BeautifulSoup(rendered_html, "html.parser")
                page_data = extract_data(soup)

                if page_data:
                    for dataset in page_data:
                        for row in dataset.get("rows", []):
                            if "Date" not in row:
                                row["Date"] = ""
                    page_results.extend(page_data)
                    print("  → JavaScript data found with Selenium")
            except Exception as e:
                print(f"  → Selenium error: {e}")

            javascript_data.extend(page_results)

    finally:
        if driver:
            driver.quit()

    return javascript_data

def remove_duplicate_datasets(data):
    unique_data = []
    seen = set()

    for dataset in data:
        headers = dataset.get("headers", [])
        rows = dataset.get("rows", [])

        # Date বাদে বাকি কন্টেন্ট দিয়ে ফিংগারপ্রিন্ট বানানো
        content_parts = []
        content_parts.append("|".join(sorted(h for h in headers if h != "Date")))

        for row in rows:
            row_values = []
            for key in sorted(row.keys()):
                if key != "Date":
                    row_values.append(f"{key}:{row.get(key, '')}")
            content_parts.append(",".join(row_values))

        fingerprint = "||".join(content_parts)

        if fingerprint not in seen:
            seen.add(fingerprint)
            unique_data.append(dataset)

    return unique_data

def merge_datasets_by_title(data):
    merged = {}

    for dataset in data:
        title = dataset.get("title", "Data Table")
        headers = dataset.get("headers", [])
        rows = dataset.get("rows", [])

        if title not in merged:
            merged[title] = {
                "title": title,
                "headers": list(headers),
                "rows": []
            }
        else:
            # নতুন হেডার এলে যোগ করা
            for h in headers:
                if h not in merged[title]["headers"]:
                    merged[title]["headers"].append(h)

        merged[title]["rows"].extend(rows)

    return list(merged.values())

def scrape_pages_with_dates(pages, date_list):
    all_data = []
    scraped_urls = set()
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 11.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"
    }

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--log-level=3")

    total_steps = len(pages) * len(date_list)
    current_step = 0

    if total_steps == 0:
        print("No pages or dates to scrape.")
        return all_data, scraped_urls

    driver = None
    try:
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        for page_url in pages:
            print(f"\n===== Page: {page_url} =====")

            for target_date in date_list:
                formatted_date = target_date.strftime("%Y-%m-%d")
                print(f"  Date: {formatted_date}")

                possible_urls = build_date_urls(page_url, formatted_date)
                found_any = False

                for target_url in possible_urls:
                    page_results = []

                    try:
                        response = requests.get(target_url, headers=headers, timeout=10)
                        if response.status_code == 200:
                            soup = BeautifulSoup(response.text, "html.parser")
                            page_data = extract_data(soup)
                            if page_data:
                                page_results.extend(page_data)
                    except requests.exceptions.RequestException:
                        pass

                    try:
                        driver.get(target_url)
                        time.sleep(2)
                        soup = BeautifulSoup(driver.page_source, "html.parser")
                        page_data = extract_data(soup)
                        if page_data:
                            page_results.extend(page_data)
                    except Exception as e:
                        print(f"    Selenium error: {e}")

                    if page_results:
                        for dataset in page_results:
                            for row in dataset.get("rows", []):
                                row["Date"] = formatted_date
                        all_data.extend(page_results)
                        scraped_urls.add(target_url.rstrip("/"))
                        found_any = True
                        print(f"    → Success")
                        break

                if not found_any:
                    print(f"    → No data")

                current_step += 1
                percent = int((current_step / total_steps) * 100)
                bar_len = 20
                filled = int(bar_len * current_step / total_steps)
                bar = "█" * filled + "-" * (bar_len - filled)
                print(f"  Progress: |{bar}| {current_step}/{total_steps} ({percent}%)")

                time.sleep(0.5)

    finally:
        if driver:
            driver.quit()

    print(f"\nScraping finished: {current_step}/{total_steps} (100%)")
    return all_data, scraped_urls

def save_to_excel(data, filename="scraped_data.xlsx"):
    if not data:
        print("No data found to save!")
        return

    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scraped Data"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Segoe UI", size=13, bold=True, color="1F4E79")
        data_font = Font(name="Segoe UI", size=10)
        row_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        row_fill_odd = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        current_row = 1
        total_rows = 0

        for dataset in data:
            title = dataset.get("title", "Data Table")
            headers = dataset.get("headers", [])
            rows = dataset.get("rows", [])

            if not headers and not rows:
                continue

            # Date সামনে আনা
            if "Date" in headers:
                headers = ["Date"] + [h for h in headers if h != "Date"]
            else:
                headers = ["Date"] + headers

            # Title লেখা (হেডারের এক লাইন উপরে)
            sheet.cell(row=current_row, column=1, value=title).font = title_font
            current_row += 1

            # হেডার লেখা
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=current_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

            # ডেটা রো
            for row_idx, row in enumerate(rows):
                fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
                for col_idx, header in enumerate(headers, start=1):
                    value = row.get(header, "")
                    cell = sheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = data_font
                    cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                current_row += 1
                total_rows += 1

            current_row += 2

        # কলাম উইডথ
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        workbook.save(filename)
        print(f"Successfully saved {total_rows} rows to '{filename}'")

    except PermissionError:
        print(f"Error: '{filename}' is open or you don't have permission to write to it. Please close the file and try again.")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

if __name__ == "__main__":
    main()
